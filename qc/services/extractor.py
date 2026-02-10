"""
POM Extractor Service (v2 - Robust)

Extracts Points of Measure (POM) data from PDF and Excel files.
Uses fuzzy column matching with RapidFuzz, header-row scoring,
split tolerance support, POM name cleaning, and noise handling
to work reliably across different customer Techfile formats
(Orsay, Softwood/SODA, generic specs, etc.).
"""

import io
import re
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass

import pdfplumber
import openpyxl
from rapidfuzz import fuzz, process


@dataclass
class ExtractedPOM:
    """Represents an extracted POM from a file"""
    name: str
    default_tol: float
    default_std: Optional[float] = None


@dataclass
class ExtractionResult:
    """Result of a file extraction operation"""
    success: bool
    poms: List[ExtractedPOM]
    matched_columns: Dict[str, str]
    confidence_scores: Dict[str, int]
    error: Optional[str] = None


# ─── Column name variations for fuzzy matching ────────────────────────────────
# Each key maps to a list of known header labels customers use.
COLUMN_PATTERNS = {
    'name': [
        'description', 'desc', 'details', 'name', 'item',
        'measurement', 'measurement point', 'measure point',
        'specification', 'spec name', 'how to measure',
        'point of measure', 'pom', 'pom description',
    ],
    'default_tol': [
        'tolerance', 'tol', '+/-', 'plus minus', 'plus/minus', 'tol (+/-)',
        'tolerance (+/-)', '+/- tol', 'allowance', 'variation', 'dev',
        'deviation', 'acceptable deviation',
    ],
    'tol_plus': [
        'tol (+)', 'tol+', 'plus', 'tol hide (+)', 'positive tolerance',
        'tolerance (+)', 'upper tolerance', 'upper tol',
    ],
    'tol_minus': [
        'tol (-)', 'tol-', 'minus', 'negative tolerance',
        'tolerance (-)', 'lower tolerance', 'lower tol',
    ],
    'default_std': [
        'standard', 'std', 'spec', 'specification', 'target', 'nominal',
        'base measurement', 'base', 'ideal', 'expected',
    ],
}

# Size headers used for scoring (not for column mapping)
SIZE_PATTERNS = [
    'xs', 's', 'm', 'l', 'xl', 'xxl', 'xxxl',
    '2xl', '3xl', '4xl',
    '34', '36', '38', '40', '42', '44', '46', '48',
    '28', '30', '32',
    'size', 'sz',
]

# Headers that indicate "noise" columns to skip entirely
NOISE_PATTERNS = ['dif', 'diff', 'difference', 'grading', 'grade']

# Minimum fuzzy match confidence threshold (0-100)
MATCH_THRESHOLD = 70


class POMExtractor:
    """Extracts POM data from PDF and Excel files"""

    # ───────────────────────────── Public API ──────────────────────────────────

    def extract(self, file_content: bytes, filename: str) -> ExtractionResult:
        """
        Extract POMs from a file based on its extension.
        
        Args:
            file_content: Raw bytes of the file
            filename: Original filename to determine file type
            
        Returns:
            ExtractionResult with extracted POMs and metadata
        """
        filename_lower = filename.lower()

        if filename_lower.endswith('.pdf'):
            return self.extract_from_pdf(file_content)
        elif filename_lower.endswith(('.xlsx', '.xls')):
            return self.extract_from_excel(file_content)
        else:
            return ExtractionResult(
                success=False, poms=[], matched_columns={},
                confidence_scores={},
                error=f"Unsupported file type: {filename}. Supported: PDF, XLSX, XLS"
            )

    # ─────────────────────────── PDF Extraction ────────────────────────────────

    def extract_from_pdf(self, file_content: bytes) -> ExtractionResult:
        """Extract POMs from a PDF file by finding tables"""
        try:
            print(f"[Extractor] Starting PDF extraction, file size: {len(file_content)} bytes")
            pdf_file = io.BytesIO(file_content)

            with pdfplumber.open(pdf_file) as pdf:
                all_tables = []
                print(f"[Extractor] PDF has {len(pdf.pages)} pages")

                # Table extraction settings – try multiple strategies
                table_settings = [
                    {},  # Default settings
                    {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
                    {"vertical_strategy": "text", "horizontal_strategy": "text"},
                    {"snap_tolerance": 5, "join_tolerance": 5},
                    {"edge_min_length": 3},  # Catches fine/dotted lines in textile TFs
                ]

                # Extract tables from all pages
                for page_idx, page in enumerate(pdf.pages):
                    print(f"[Extractor] Processing page {page_idx + 1}")

                    # Try different extraction strategies
                    for strategy_idx, settings in enumerate(table_settings):
                        try:
                            tables = page.extract_tables(table_settings=settings)
                            if tables:
                                print(f"[Extractor] Page {page_idx + 1}, strategy {strategy_idx}: found {len(tables)} tables")
                                for table in tables:
                                    if table and len(table) > 1:
                                        print(f"[Extractor] Found table with {len(table)} rows")
                                        if table[0]:
                                            print(f"[Extractor] Table row 0: {table[0][:5]}...")
                                        if len(table) > 1 and table[1]:
                                            print(f"[Extractor] Table row 1: {table[1][:5]}...")
                                        if len(table) > 2 and table[2]:
                                            print(f"[Extractor] Table row 2: {table[2][:5]}...")
                                        all_tables.append(table)
                                if all_tables:
                                    break  # Use first successful strategy
                        except Exception as e:
                            print(f"[Extractor] Strategy {strategy_idx} failed: {e}")
                            continue

                if not all_tables:
                    # ── Fallback: try text-based extraction ───────────────
                    print("[Extractor] No tables found, trying text-based fallback")
                    result = self._extract_from_text(pdf)
                    if result.success:
                        return result
                    return ExtractionResult(
                        success=False, poms=[], matched_columns={},
                        confidence_scores={},
                        error="No tables found in PDF. Please ensure the PDF contains a measurement table."
                    )

                print(f"[Extractor] Total tables found: {len(all_tables)}")

                # Find the best table (one with matching columns)
                for table in all_tables:
                    result = self._process_table(table)
                    if result.success:
                        print(f"[Extractor] Successfully extracted {len(result.poms)} POMs")
                        return result

                # If no table matched, try the largest one
                largest_table = max(all_tables, key=lambda t: len(t))
                print(f"[Extractor] No direct match, trying largest table ({len(largest_table)} rows)")
                return self._process_table(largest_table)

        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"[Extractor] Exception during PDF extraction:\n{error_trace}")
            return ExtractionResult(
                success=False, poms=[], matched_columns={},
                confidence_scores={},
                error=f"Error reading PDF: {str(e)}"
            )

    # ─────────────────────────── Excel Extraction ──────────────────────────────

    def extract_from_excel(self, file_content: bytes) -> ExtractionResult:
        """Extract POMs from an Excel file"""
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                table = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        table.append([str(cell) if cell is not None else '' for cell in row])

                if len(table) > 1:
                    result = self._process_table(table)
                    if result.success and result.poms:
                        return result

            return ExtractionResult(
                success=False, poms=[], matched_columns={},
                confidence_scores={},
                error="No measurement data found in Excel file. Please ensure it contains a table with POM and Tolerance columns."
            )
        except Exception as e:
            return ExtractionResult(
                success=False, poms=[], matched_columns={},
                confidence_scores={},
                error=f"Error reading Excel file: {str(e)}"
            )

    # ─────────────────── Text-Based Fallback (no tables) ───────────────────────

    def _extract_from_text(self, pdf) -> ExtractionResult:
        """
        Fallback: extract POM data from raw text when pdfplumber
        detects no table structure.
        Looks for lines matching: <text description> <number> <optional tolerance>
        """
        poms = []
        # Pattern: text description, then one or more numbers
        # e.g. "Shoulder Width  15.5  +/-0.5"  or  "Waist  80  1.0"
        line_pattern = re.compile(
            r'^([A-Za-z][A-Za-z\s/\-&().]{2,}?)'  # name (starts with letter, >=3 chars)
            r'\s+'
            r'([\d]+\.?\d*)'                       # first number (std or ignored)
            r'(?:\s+[±+\-/]*\s*([\d]+\.?\d*))?'    # optional tolerance number
        )

        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.split('\n'):
                line = line.strip()
                if not line:
                    continue
                m = line_pattern.match(line)
                if m:
                    name = self._clean_pom_name(m.group(1).strip())
                    if len(name) < 3:
                        continue
                    tol_str = m.group(3) or '0'
                    default_tol = abs(self._parse_number(tol_str))
                    poms.append(ExtractedPOM(name=name, default_tol=default_tol))

        if poms:
            print(f"[Extractor] Text fallback extracted {len(poms)} POMs")
            return ExtractionResult(
                success=True, poms=poms,
                matched_columns={'name': 'text', 'default_tol': 'text'},
                confidence_scores={'name': 60, 'default_tol': 60}
            )

        return ExtractionResult(
            success=False, poms=[], matched_columns={},
            confidence_scores={},
            error="Text fallback found no measurement-like lines."
        )

    # ─────────────────────────── Table Processing ──────────────────────────────

    def _process_table(self, table: List[List[str]]) -> ExtractionResult:
        """
        Process a table to extract POM data.
        Uses a scoring system to find the best header row, then
        extracts data rows beneath it.
        """
        try:
            if not table or len(table) < 2:
                return ExtractionResult(
                    success=False, poms=[], matched_columns={},
                    confidence_scores={},
                    error="Table too small (needs header + at least 1 data row)"
                )

            # ── Step 1: Normalize all cells (multi-newlines → single space) ───
            table = self._normalize_table(table)

            # ── Step 2: Find the best header row via scoring ──────────────────
            header_row, data_start_idx = self._find_header_row(table)

            if not header_row:
                return ExtractionResult(
                    success=False, poms=[], matched_columns={},
                    confidence_scores={},
                    error="Could not find header row in table"
                )

            # ── Step 3: Match columns using fuzzy matching ────────────────────
            column_mapping, confidence_scores = self._match_columns(header_row)
            print(f"[Extractor] Column mapping result: {column_mapping}")
            print(f"[Extractor] Confidence scores: {confidence_scores}")

            # We need at least the name column
            if 'name' not in column_mapping:
                # ── Step 3b: Try data-type inference as last resort ────────
                column_mapping, confidence_scores = self._infer_columns_from_data(
                    table, data_start_idx, header_row
                )
                if 'name' not in column_mapping:
                    non_empty_cols = [col for col in header_row if col]
                    return ExtractionResult(
                        success=False, poms=[], matched_columns={},
                        confidence_scores={},
                        error=f"Could not find a Description/POM name column. Available columns: {', '.join(non_empty_cols)}"
                    )

            # ── Step 4: Identify noise columns to skip ────────────────────────
            noise_col_indices = self._find_noise_columns(header_row)

            matched_columns = {
                field: header_row[col_idx]
                for field, col_idx in column_mapping.items()
                if col_idx < len(header_row)
            }

            # ── Step 5: Extract data rows ─────────────────────────────────────
            poms = self._extract_rows(
                table, data_start_idx, column_mapping,
                header_row, noise_col_indices
            )

            if not poms:
                return ExtractionResult(
                    success=False, poms=[], matched_columns=matched_columns,
                    confidence_scores=confidence_scores,
                    error="No valid POM data found in table rows"
                )

            return ExtractionResult(
                success=True, poms=poms,
                matched_columns=matched_columns,
                confidence_scores=confidence_scores
            )
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"[Extractor] Exception in _process_table:\n{error_trace}")
            return ExtractionResult(
                success=False, poms=[], matched_columns={},
                confidence_scores={},
                error=f"Error processing table: {str(e)}"
            )

    # ──────────────────────── Header Scoring System ────────────────────────────

    def _find_header_row(self, table: List[List[str]]) -> Tuple[Optional[List[str]], int]:
        """
        Scan the first 10 rows and score each one.
        Scoring:
          +2  if a cell matches a 'name' pattern
          +1  if a cell matches 'default_tol', 'tol_plus', or 'tol_minus'
          +1  if a cell matches a common size header
        The row with the highest score wins.
        """
        best_row = None
        best_score = -1
        best_idx = 0
        scan_limit = min(10, len(table))

        for idx in range(scan_limit):
            row = table[idx]
            if not any(cell and str(cell).strip() for cell in row):
                continue

            potential_header = [str(cell).strip() if cell else '' for cell in row]
            score = self._score_header_row(potential_header)

            print(f"[Extractor] Row {idx} score={score}: {potential_header[:6]}...")

            if score > best_score:
                best_score = score
                best_row = potential_header
                best_idx = idx

        if best_row and best_score >= 2:
            print(f"[Extractor] Best header row at index {best_idx} (score={best_score})")
            return best_row, best_idx + 1

        # Fallback: first non-empty row
        for idx, row in enumerate(table):
            if any(cell and str(cell).strip() for cell in row):
                fallback = [str(cell).strip() if cell else '' for cell in row]
                print(f"[Extractor] Using fallback: first non-empty row at index {idx}")
                return fallback, idx + 1

        return None, 0

    def _score_header_row(self, row: List[str]) -> int:
        """Score a candidate header row for likelihood of being the actual header."""
        score = 0
        for cell in row:
            if not cell:
                continue
            cell_lower = cell.lower().strip()

            # +2 for 'name' patterns
            if self._cell_matches_patterns(cell_lower, COLUMN_PATTERNS['name']):
                score += 2

            # +1 for tolerance patterns (unified, plus, minus)
            for tol_field in ('default_tol', 'tol_plus', 'tol_minus'):
                if tol_field in COLUMN_PATTERNS and self._cell_matches_patterns(cell_lower, COLUMN_PATTERNS[tol_field]):
                    score += 1

            # +1 for size patterns
            if cell_lower in SIZE_PATTERNS:
                score += 1

            # +1 for 'default_std' patterns
            if self._cell_matches_patterns(cell_lower, COLUMN_PATTERNS['default_std']):
                score += 1

        return score

    def _cell_matches_patterns(self, cell_lower: str, patterns: List[str]) -> bool:
        """Check if a cell matches any of the given patterns (exact or fuzzy)."""
        if cell_lower in patterns:
            return True
        for pattern in patterns:
            score = max(fuzz.ratio(cell_lower, pattern), fuzz.partial_ratio(cell_lower, pattern))
            if score >= MATCH_THRESHOLD:
                return True
        return False

    # ──────────────────────── Column Matching ──────────────────────────────────

    # Process fields in this order so 'default_tol' claims 'Tol (+/-)'
    # before 'tol_plus' can grab it via partial match.
    FIELD_PRIORITY = ['name', 'default_tol', 'default_std', 'tol_plus', 'tol_minus']

    def _match_columns(self, headers: List[str]) -> Tuple[Dict[str, int], Dict[str, int]]:
        """
        Match header columns to expected fields using fuzzy matching.
        Uses a TWO-PASS approach:
          Pass 1: Exact matches only (all fields) – prevents fuzzy cross-matching
          Pass 2: Fuzzy matches on remaining unmatched fields/columns

        Returns:
            Tuple of (column_mapping, confidence_scores)
        """
        column_mapping = {}
        confidence_scores = {}
        used_indices = set()

        # Prepare cleaned headers once
        cleaned_headers = []
        for idx, header in enumerate(headers):
            h = header.lower().strip() if header else ''
            cleaned_headers.append(h)

        # ── Pass 1: Exact matches for all fields ─────────────────────
        for field in self.FIELD_PRIORITY:
            if field not in COLUMN_PATTERNS:
                continue
            patterns = COLUMN_PATTERNS[field]

            for idx, header_lower in enumerate(cleaned_headers):
                if not header_lower or idx in used_indices:
                    continue
                if len(header_lower) < 2:
                    continue
                if self._is_noise_column(header_lower):
                    continue

                if header_lower in patterns:
                    column_mapping[field] = idx
                    confidence_scores[field] = 100
                    used_indices.add(idx)
                    break

        # ── Pass 2: Fuzzy matches for remaining unmatched fields ─────
        for field in self.FIELD_PRIORITY:
            if field in column_mapping:  # Already exact-matched
                continue
            if field not in COLUMN_PATTERNS:
                continue
            patterns = COLUMN_PATTERNS[field]
            best_match_idx = None
            best_score = 0

            # For split-tol fields, use strict ratio and higher threshold.
            use_strict = field in ('tol_plus', 'tol_minus')
            threshold = 85 if use_strict else MATCH_THRESHOLD

            for idx, header_lower in enumerate(cleaned_headers):
                if not header_lower or idx in used_indices:
                    continue
                if len(header_lower) < 3:
                    continue
                if self._is_noise_column(header_lower):
                    continue

                for pattern in patterns:
                    if use_strict:
                        score = fuzz.ratio(header_lower, pattern)
                    else:
                        score = max(
                            fuzz.ratio(header_lower, pattern),
                            fuzz.partial_ratio(header_lower, pattern)
                        )
                    if score > best_score and score >= threshold:
                        best_score = score
                        best_match_idx = idx

            if best_match_idx is not None:
                column_mapping[field] = best_match_idx
                confidence_scores[field] = best_score
                used_indices.add(best_match_idx)

        return column_mapping, confidence_scores

    # ───────────────────── Data-Type Inference (fallback) ──────────────────────

    def _infer_columns_from_data(
        self, table: List[List[str]], data_start: int,
        header_row: List[str]
    ) -> Tuple[Dict[str, int], Dict[str, int]]:
        """
        When header matching fails, guess column types by analysing
        the content of data rows.
          - Column with >60% text-like values → 'name'
          - Column with >60% small numbers → 'default_tol'
          - Column with >60% larger numbers → 'default_std'
        """
        print("[Extractor] Attempting data-type inference for columns")
        data_rows = table[data_start: data_start + 20]  # Sample up to 20 rows
        if not data_rows:
            return {}, {}

        num_cols = max(len(r) for r in data_rows)
        col_stats = []  # (text_pct, numeric_pct, avg_value)

        for col_idx in range(num_cols):
            text_count = 0
            num_count = 0
            num_sum = 0.0
            total = 0

            # Skip noise columns
            if col_idx < len(header_row) and self._is_noise_column(header_row[col_idx].lower()):
                col_stats.append((0, 0, 0))
                continue

            for row in data_rows:
                if col_idx >= len(row):
                    continue
                cell = str(row[col_idx]).strip() if row[col_idx] else ''
                if not cell:
                    continue
                total += 1
                val = self._parse_number(cell)
                if val != 0.0 or cell in ('0', '0.0', '0,0'):
                    num_count += 1
                    num_sum += abs(val)
                else:
                    # Check if it's truly text (contains letters)
                    if re.search(r'[a-zA-Z]', cell):
                        text_count += 1

            if total == 0:
                col_stats.append((0, 0, 0))
            else:
                text_pct = text_count / total
                num_pct = num_count / total
                avg_val = num_sum / num_count if num_count else 0
                col_stats.append((text_pct, num_pct, avg_val))

        # Find best column for each role
        column_mapping = {}
        confidence_scores = {}
        used = set()

        # 'name': highest text_pct (>0.6)
        best_name_idx = -1
        best_name_pct = 0
        for i, (tp, np, av) in enumerate(col_stats):
            if i in used:
                continue
            if tp > best_name_pct and tp > 0.6:
                best_name_pct = tp
                best_name_idx = i
        if best_name_idx >= 0:
            column_mapping['name'] = best_name_idx
            confidence_scores['name'] = int(best_name_pct * 100)
            used.add(best_name_idx)
            print(f"[Extractor] Inferred 'name' at col {best_name_idx} ({best_name_pct:.0%} text)")

        # 'default_tol': numeric column with smallest avg value (tolerance ≈ small numbers)
        best_tol_idx = -1
        best_tol_avg = float('inf')
        for i, (tp, np, av) in enumerate(col_stats):
            if i in used:
                continue
            if np > 0.6 and av < best_tol_avg:
                best_tol_avg = av
                best_tol_idx = i
        if best_tol_idx >= 0:
            column_mapping['default_tol'] = best_tol_idx
            confidence_scores['default_tol'] = int(col_stats[best_tol_idx][1] * 100)
            used.add(best_tol_idx)
            print(f"[Extractor] Inferred 'default_tol' at col {best_tol_idx} (avg={best_tol_avg:.2f})")

        return column_mapping, confidence_scores

    # ───────────────────────── Noise Detection ─────────────────────────────────

    def _is_noise_column(self, header_lower: str) -> bool:
        """Check if a column header indicates noise (DIF, DIFF, etc.)."""
        for noise in NOISE_PATTERNS:
            if noise in header_lower:
                return True
        return False

    def _find_noise_columns(self, header_row: List[str]) -> set:
        """Return set of column indices that are noise columns."""
        noise = set()
        for idx, h in enumerate(header_row):
            if h and self._is_noise_column(h.lower().strip()):
                noise.add(idx)
        if noise:
            print(f"[Extractor] Noise columns detected at indices: {noise}")
        return noise

    # ─────────────────────── Row Extraction Logic ──────────────────────────────

    def _extract_rows(
        self, table: List[List[str]], data_start: int,
        column_mapping: Dict[str, int], header_row: List[str],
        noise_cols: set
    ) -> List[ExtractedPOM]:
        """Extract POM objects from data rows beneath the header."""
        poms = []
        has_split_tol = 'tol_plus' in column_mapping or 'tol_minus' in column_mapping

        for row_idx, row in enumerate(table[data_start:]):
            if not row or not any(str(cell).strip() if cell else '' for cell in row):
                continue

            # Pad row if needed
            while len(row) < len(header_row):
                row.append('')

            # ── Get raw name ──────────────────────────────────────────────
            name_idx = column_mapping['name']
            raw_name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] else ''
            if not raw_name:
                continue

            # Handle newlines in name cells (merged cells)
            # Split on actual newlines or literal escaped \n from PDF extraction
            names = [n.strip() for n in re.split(r'\n|\\n', raw_name) if n.strip()]

            # ── Get tolerance ─────────────────────────────────────────────
            default_tol = 0.0
            if has_split_tol:
                default_tol = self._get_split_tolerance(row, column_mapping)
            else:
                tol_idx = column_mapping.get('default_tol')
                if tol_idx is not None and tol_idx < len(row) and row[tol_idx]:
                    raw_tol = str(row[tol_idx]).strip()
                    tol_values = [t.strip() for t in re.split(r'\n|\\n', raw_tol) if t.strip()]
                    if tol_values:
                        default_tol = abs(self._parse_number(tol_values[0]))

            # ── Get standard (optional) ───────────────────────────────────
            default_std = None
            if 'default_std' in column_mapping:
                std_idx = column_mapping['default_std']
                if std_idx < len(row) and row[std_idx]:
                    std_value = self._parse_number(str(row[std_idx]))
                    if std_value != 0.0:
                        default_std = std_value

            # ── Create POM(s) ─────────────────────────────────────────────
            if len(names) > 1:
                # Multiple names in one cell → one POM per name
                for name in names:
                    cleaned = self._clean_pom_name(name)
                    if cleaned:
                        poms.append(ExtractedPOM(
                            name=cleaned,
                            default_tol=default_tol,
                            default_std=None
                        ))
            else:
                cleaned = self._clean_pom_name(names[0] if names else raw_name)
                if cleaned:
                    poms.append(ExtractedPOM(
                        name=cleaned,
                        default_tol=default_tol,
                        default_std=default_std
                    ))

        return poms

    # ────────────────── Split Tolerance (plus / minus) ─────────────────────────

    def _get_split_tolerance(self, row: List[str], column_mapping: Dict[str, int]) -> float:
        """
        Combine tol_plus and tol_minus into a single tolerance value.
        Uses max(abs(plus), abs(minus)).
        Falls back to default_tol if split columns are missing.
        """
        plus_val = 0.0
        minus_val = 0.0

        tol_plus_idx = column_mapping.get('tol_plus')
        if tol_plus_idx is not None and tol_plus_idx < len(row) and row[tol_plus_idx]:
            plus_val = abs(self._parse_number(str(row[tol_plus_idx])))

        tol_minus_idx = column_mapping.get('tol_minus')
        if tol_minus_idx is not None and tol_minus_idx < len(row) and row[tol_minus_idx]:
            minus_val = abs(self._parse_number(str(row[tol_minus_idx])))

        combined = max(plus_val, minus_val)

        # If split tol is 0, fallback to unified tol column
        if combined == 0.0:
            tol_idx = column_mapping.get('default_tol')
            if tol_idx is not None and tol_idx < len(row) and row[tol_idx]:
                combined = abs(self._parse_number(str(row[tol_idx])))

        return combined

    # ─────────────────────── POM Name Cleaning ─────────────────────────────────

    def _clean_pom_name(self, raw_name: str) -> str:
        """
        Clean a POM name by removing leading codes/numbers.
        Examples:
            "42 A Waist width"     → "Waist width"
            "Dim H1 Shoulder"      → "Shoulder"
            "A. Back length"       → "Back length"
            "1. Chest width"       → "Chest width"
            "H1 Front rise"        → "Front rise"
            "Shoulder width"       → "Shoulder width"  (unchanged)
        """
        if not raw_name:
            return ''

        name = raw_name.strip()

        # Strategy: try several specific patterns in order.
        # Each pattern strips a known code prefix format.

        # Pattern 1: "Dim H1 " prefix
        cleaned = re.sub(r'^Dim\s+[A-Z0-9]+\s+', '', name)
        if cleaned != name:
            return cleaned.strip(' .-') or name

        # Pattern 2: "42 A " → number + space + single letter + space
        cleaned = re.sub(r'^\d+\s+[A-Z]\s+', '', name)
        if cleaned != name:
            return cleaned.strip(' .-') or name

        # Pattern 3: "A. ", "B2. ", "1. " → short code + dot + space
        cleaned = re.sub(r'^[A-Z0-9]{1,3}\.\s*', '', name)
        if cleaned != name:
            return cleaned.strip(' .-') or name

        # Pattern 4: "H1 ", "A " → 1-2 char uppercase code + space,
        # but ONLY if followed by a capitalized word (to avoid stripping
        # real words like "Front" from "Front rise")
        cleaned = re.sub(r'^[A-Z][0-9]\s+(?=[A-Z][a-z])', '', name)
        if cleaned != name:
            return cleaned.strip(' .-') or name

        # Pattern 5: Leading bare number + space (e.g. "3 Front rise")
        cleaned = re.sub(r'^\d+\s+', '', name)
        if cleaned != name:
            return cleaned.strip(' .-') or name

        # No code prefix found — return as-is
        return name.strip(' .-')

    # ───────────────────── Table Normalization ─────────────────────────────────

    def _normalize_table(self, table: List[List[str]]) -> List[List[str]]:
        """
        Normalize all cells in a table:
          - Replace multiple newlines with a single space
          - Strip whitespace
          - Convert None to empty string
        """
        normalized = []
        for row in table:
            new_row = []
            for cell in row:
                if cell is None:
                    new_row.append('')
                else:
                    # Replace multiple newlines (actual \n) with single space
                    s = str(cell)
                    s = re.sub(r'\n+', ' ', s)
                    # Also handle literal escaped newlines from PDF extraction
                    s = re.sub(r'(\\n)+', ' ', s)
                    new_row.append(s.strip())
            normalized.append(new_row)
        return normalized

    # ───────────────────── Number Parsing ──────────────────────────────────────

    def _parse_number(self, value: str) -> float:
        """
        Parse a string to a float, handling various formats:
          - European decimals: "1,5" → 1.5
          - +/- prefix: "+/-1.0" → 1.0
          - Units suffix: "1.5 cm" → 1.5
          - Fractions: "1/2" → 0.5, "3/4" → 0.75
        """
        if not value:
            return 0.0

        cleaned = value.strip()
        cleaned = cleaned.replace(',', '.')  # European decimal format

        # Remove +/- prefix
        if cleaned.startswith('+/-') or cleaned.startswith('±'):
            cleaned = cleaned[3:] if cleaned.startswith('+/-') else cleaned[1:]
        
        # Remove unit suffixes
        cleaned = re.sub(r'\s*(cm|mm|in|inch|inches|")\s*$', '', cleaned, flags=re.IGNORECASE)

        # Try fraction first (e.g. "1/2", "3/4")
        fraction_match = re.match(r'^[-+]?\s*(\d+)\s*/\s*(\d+)$', cleaned)
        if fraction_match:
            try:
                num = float(fraction_match.group(1))
                den = float(fraction_match.group(2))
                return num / den if den != 0 else 0.0
            except (ValueError, ZeroDivisionError):
                pass

        # Extract numeric part
        match = re.search(r'[-+]?\d*\.?\d+', cleaned)
        if match:
            try:
                return float(match.group())
            except ValueError:
                return 0.0

        return 0.0


# ─────────────────────── Convenience Function ─────────────────────────────────

def extract_pom_from_file(file_content: bytes, filename: str) -> ExtractionResult:
    """
    Extract POMs from a file.

    Args:
        file_content: Raw bytes of the file
        filename: Original filename to determine file type

    Returns:
        ExtractionResult with extracted POMs and metadata
    """
    extractor = POMExtractor()
    return extractor.extract(file_content, filename)
